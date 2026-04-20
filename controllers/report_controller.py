"""
报表控制器
"""
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from datetime import datetime
from dateutil.parser import parse
import os
import uuid
from core.database import get_db
from models.account import Account, Transaction
from models.user import User
from schemas import ReportRequest, ReportResponse, ApiResponse
from core.auth import require_login
from core import config

router = APIRouter(prefix="/api/reports", tags=["报表"])

os.makedirs(config.REPORT_DIR, exist_ok=True)


@router.post("/generate")
async def generate_report(
    request: Request,
    report_data: ReportRequest,
    db: Session = Depends(get_db)
):
    """生成报表"""
    user = await require_login(request, db)
    
    try:
        if report_data.report_type == "daily":
            file_path = generate_daily_report(db, user, report_data)
        elif report_data.report_type == "monthly":
            file_path = generate_monthly_report(db, user, report_data)
        elif report_data.report_type == "account":
            file_path = generate_account_report(db, user, report_data)
        else:
            return ApiResponse.error(message="无效的报表类型", code=400)
        
        return ApiResponse.success(
            data=ReportResponse(success=True, message="报表生成成功", file_path=file_path),
            message="报表生成成功"
        )
    except Exception as e:
        return ApiResponse.error(message=f"报表生成失败: {str(e)}")


def generate_daily_report(db: Session, user: User, report_data: ReportRequest) -> str:
    """生成日报"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    date_str = report_data.start_date or datetime.now().strftime("%Y-%m-%d")
    report_date = parse(date_str).date()
    
    start_datetime = datetime.combine(report_date, datetime.min.time())
    end_datetime = datetime.combine(report_date, datetime.max.time())
    
    query = db.query(Transaction).join(Account).filter(
        Transaction.created_at >= start_datetime,
        Transaction.created_at <= end_datetime,
        Account.owner_id == user.id
    ).order_by(Transaction.created_at)
    
    transactions = query.all()
    
    file_name = f"daily_report_{report_date}_{uuid.uuid4().hex[:8]}.pdf"
    file_path = os.path.join(config.REPORT_DIR, file_name)
    
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph(f"{config.BANK_NAME} 日报", styles['Title']))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"报表日期: {report_date}", styles['Normal']))
    elements.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    total_deposit = sum(float(t.amount) for t in transactions if t.transaction_type == "deposit")
    total_withdraw = sum(float(t.amount) for t in transactions if t.transaction_type == "withdraw")
    total_transfer_in = sum(float(t.amount) for t in transactions if t.transaction_type == "transfer_in")
    total_transfer_out = sum(float(t.amount) for t in transactions if t.transaction_type == "transfer_out")
    
    elements.append(Paragraph(f"交易笔数: {len(transactions)}", styles['Normal']))
    elements.append(Paragraph(f"存款总额: ¥{total_deposit:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"取款总额: ¥{total_withdraw:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"转入总额: ¥{total_transfer_in:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"转出总额: ¥{total_transfer_out:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    if transactions:
        data = [["时间", "类型", "金额", "余额", "描述"]]
        type_map = {"deposit": "存款", "withdraw": "取款", "transfer_in": "转入", "transfer_out": "转出"}
        for t in transactions:
            data.append([
                t.created_at.strftime("%H:%M:%S"),
                type_map.get(t.transaction_type, t.transaction_type),
                f"¥{float(t.amount):,.2f}",
                f"¥{float(t.balance_after):,.2f}",
                t.description or "-"
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
    
    doc.build(elements)
    return file_path


def generate_monthly_report(db: Session, user: User, report_data: ReportRequest) -> str:
    """生成月报"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    date_str = report_data.start_date or datetime.now().strftime("%Y-%m")
    report_month = parse(date_str).date().replace(day=1)
    
    start_datetime = datetime.combine(report_month, datetime.min.time())
    if report_month.month == 12:
        end_datetime = datetime.combine(report_month.replace(year=report_month.year + 1, month=1), datetime.min.time())
    else:
        end_datetime = datetime.combine(report_month.replace(month=report_month.month + 1), datetime.min.time())
    
    query = db.query(Transaction).join(Account).filter(
        Transaction.created_at >= start_datetime,
        Transaction.created_at < end_datetime,
        Account.owner_id == user.id
    )
    
    transactions = query.all()
    
    daily_stats = {}
    for t in transactions:
        day = t.created_at.date()
        if day not in daily_stats:
            daily_stats[day] = {"count": 0, "deposit": 0, "withdraw": 0, "transfer_in": 0, "transfer_out": 0}
        daily_stats[day]["count"] += 1
        if t.transaction_type == "deposit":
            daily_stats[day]["deposit"] += float(t.amount)
        elif t.transaction_type == "withdraw":
            daily_stats[day]["withdraw"] += float(t.amount)
        elif t.transaction_type == "transfer_in":
            daily_stats[day]["transfer_in"] += float(t.amount)
        elif t.transaction_type == "transfer_out":
            daily_stats[day]["transfer_out"] += float(t.amount)
    
    file_name = f"monthly_report_{report_month}_{uuid.uuid4().hex[:8]}.pdf"
    file_path = os.path.join(config.REPORT_DIR, file_name)
    
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph(f"{config.BANK_NAME} 月报", styles['Title']))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"报表月份: {report_month.strftime('%Y-%m')}", styles['Normal']))
    elements.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    total_deposit = sum(float(t.amount) for t in transactions if t.transaction_type == "deposit")
    total_withdraw = sum(float(t.amount) for t in transactions if t.transaction_type == "withdraw")
    
    elements.append(Paragraph(f"交易总笔数: {len(transactions)}", styles['Normal']))
    elements.append(Paragraph(f"存款总额: ¥{total_deposit:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"取款总额: ¥{total_withdraw:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    if daily_stats:
        data = [["日期", "笔数", "存款", "取款", "转入", "转出"]]
        for day in sorted(daily_stats.keys()):
            stats = daily_stats[day]
            data.append([
                day.strftime("%m-%d"),
                str(stats["count"]),
                f"¥{stats['deposit']:,.2f}",
                f"¥{stats['withdraw']:,.2f}",
                f"¥{stats['transfer_in']:,.2f}",
                f"¥{stats['transfer_out']:,.2f}"
            ])
        
        table = Table(data, colWidths=[80, 50, 80, 80, 80, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
    
    doc.build(elements)
    return file_path


def generate_account_report(db: Session, user: User, report_data: ReportRequest) -> str:
    """生成账户报表"""
    if report_data.format == "excel":
        return generate_account_excel(db, user, report_data)
    else:
        return generate_account_pdf(db, user, report_data)


def generate_account_pdf(db: Session, user: User, report_data: ReportRequest) -> str:
    """生成账户PDF报表"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    account_number = report_data.account_number
    if not account_number:
        raise ValueError("账户号不能为空")
    
    account = db.query(Account).filter(
        Account.account_number == account_number,
        Account.owner_id == user.id
    ).first()
    
    if not account:
        raise ValueError("账户不存在")
    
    start_date = datetime.combine(parse(report_data.start_date).date(), datetime.min.time()) if report_data.start_date else datetime(2000, 1, 1)
    end_date = datetime.combine(parse(report_data.end_date).date(), datetime.max.time()) if report_data.end_date else datetime.now()
    
    transactions = db.query(Transaction).filter(
        Transaction.account_id == account.id,
        Transaction.created_at >= start_date,
        Transaction.created_at <= end_date
    ).order_by(Transaction.created_at.desc()).limit(100).all()
    
    file_name = f"account_report_{account_number}_{uuid.uuid4().hex[:8]}.pdf"
    file_path = os.path.join(config.REPORT_DIR, file_name)
    
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph(f"{config.BANK_NAME} 账户报表", styles['Title']))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"账户号: {account_number}", styles['Normal']))
    elements.append(Paragraph(f"账户类型: {'储蓄账户' if account.account_type == 'savings' else '支票账户'}", styles['Normal']))
    elements.append(Paragraph(f"当前余额: ¥{float(account.balance):,.2f}", styles['Normal']))
    elements.append(Paragraph(f"账户状态: {'正常' if account.status == 'active' else '冻结' if account.status == 'frozen' else '已关闭'}", styles['Normal']))
    elements.append(Paragraph(f"报表时间: {start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    if transactions:
        data = [["时间", "类型", "金额", "余额", "描述"]]
        type_map = {"deposit": "存款", "withdraw": "取款", "transfer_in": "转入", "transfer_out": "转出"}
        for t in transactions:
            data.append([
                t.created_at.strftime("%Y-%m-%d %H:%M"),
                type_map.get(t.transaction_type, t.transaction_type),
                f"¥{float(t.amount):,.2f}",
                f"¥{float(t.balance_after):,.2f}",
                (t.description or "-")[:20]
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("暂无交易记录", styles['Normal']))
    
    doc.build(elements)
    return file_path


def generate_account_excel(db: Session, user: User, report_data: ReportRequest) -> str:
    """生成账户Excel报表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    
    account_number = report_data.account_number
    if not account_number:
        raise ValueError("账户号不能为空")
    
    account = db.query(Account).filter(
        Account.account_number == account_number,
        Account.owner_id == user.id
    ).first()
    
    if not account:
        raise ValueError("账户不存在")
    
    start_date = datetime.combine(parse(report_data.start_date).date(), datetime.min.time()) if report_data.start_date else datetime(2000, 1, 1)
    end_date = datetime.combine(parse(report_data.end_date).date(), datetime.max.time()) if report_data.end_date else datetime.now()
    
    transactions = db.query(Transaction).filter(
        Transaction.account_id == account.id,
        Transaction.created_at >= start_date,
        Transaction.created_at <= end_date
    ).order_by(Transaction.created_at.desc()).limit(100).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "交易记录"
    
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["时间", "类型", "金额", "余额", "描述"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    type_map = {"deposit": "存款", "withdraw": "取款", "transfer_in": "转入", "transfer_out": "转出"}
    
    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=t.created_at.strftime("%Y-%m-%d %H:%M")).border = thin_border
        ws.cell(row=row, column=2, value=type_map.get(t.transaction_type, t.transaction_type)).border = thin_border
        ws.cell(row=row, column=3, value=float(t.amount)).border = thin_border
        ws.cell(row=row, column=4, value=float(t.balance_after)).border = thin_border
        ws.cell(row=row, column=5, value=t.description or "-").border = thin_border
    
    file_name = f"account_report_{account_number}_{uuid.uuid4().hex[:8]}.xlsx"
    file_path = os.path.join(config.REPORT_DIR, file_name)
    wb.save(file_path)
    
    return file_path


@router.get("/download/{filename}")
async def download_report(
    request: Request,
    filename: str,
    db: Session = Depends(get_db)
):
    """下载报表"""
    user = await require_login(request, db)
    
    file_path = os.path.join(config.REPORT_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="文件不存在")
    
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type='application/octet-stream'
    )


@router.get("/list")
async def list_reports(
    request: Request,
    db: Session = Depends(get_db)
):
    """获取报表列表"""
    user = await require_login(request, db)
    
    files = []
    for f in os.listdir(config.REPORT_DIR):
        file_path = os.path.join(config.REPORT_DIR, f)
        if os.path.isfile(file_path):
            stat = os.stat(file_path)
            files.append({
                "name": f,
                "size": stat.st_size,
                "created": datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M:%S")
            })
    
    return ApiResponse.success(
        data=sorted(files, key=lambda x: x["created"], reverse=True),
        message="查询成功"
    )
